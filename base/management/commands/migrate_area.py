import openpyxl

from django.core.management.base import BaseCommand

from base.models import Area


class Command(BaseCommand):
    """
    지역 마이그레이션 커맨드
    """
    start_row = 4

    def handle(self, *args, **options):
        data = self.handle_excel()

    def handle_excel(self):
        wb = openpyxl.load_workbook('web/migration_data/dong.xlsx')
        provinces_count, cities_count = self.create_addresses(wb[wb.sheetnames[0]])
        self.stdout.write(self.style.SUCCESS('%s개 시/도와 %s개 구/군이 추가되었습니다.' % (provinces_count, cities_count)))
        conntected_count = self.connect_addresses(wb[wb.sheetnames[1]])
        self.stdout.write(self.style.SUCCESS('%s개 인근지역이 추가되었습니다.' % conntected_count))

    def create_addresses(self, sheet):
        province_text = ''
        province_obj = None
        city_text = ''
        city_obj = None

        provinces_count = 0
        cities_count = 0

        for row in list(sheet)[self.start_row-1:]:
            val1 = row[0].value.strip()
            if province_text != val1:
                province_text = val1
                province_obj, created = Area.objects.get_or_create(name=province_text, parent=None)
                if created:
                    provinces_count += 1

        for row in list(sheet)[self.start_row-1:]:
            val = [r.value.strip() if r.value else None for r in row]
            if val[0] and val[1]:
                if province_text != val[0]:
                    province_text = val[0]
                    province_obj = Area.objects.get(name=province_text, parent=None)
                if city_text != val[1]:
                    city_text = val[1]
                    city_obj, created = Area.objects.get_or_create(name=city_text, parent=province_obj)
                    if created:
                        cities_count += 1

        return provinces_count, cities_count

    def connect_addresses(self, sheet):
        obj = None
        connected_count = 0

        for row in list(sheet)[self.start_row-1:]:

            # 빈 row가 나오면 중단
            if not (row[0].value or row[1].value or row[2].value or row[3].value):
                break

            # 셀 텍스트에 맞는 오브젝트 찾아오기
            val = [r.value.strip() if r.value else None for r in row]
            if not obj or (obj.parent and obj.parent.name != val[0]) or obj.name != val[1]:
                if val[1]:
                    obj = Area.objects.filter(name=val[1], parent__name=val[0]).last()
                else:
                    obj = Area.objects.filter(name=val[0], parent__isnull=True).last()
            if val[3]:
                target = Area.objects.filter(name=val[3], parent__name=val[2]).last()
            else:
                if val[2]:
                    target = Area.objects.filter(name=val[2], parent__isnull=True).last()
                else:
                    continue

            # 저장
            if obj and target:
                if target not in obj.nearby.all():
                    obj.nearby.add(target)
                    obj.save()
                    connected_count += 1
            else:
                self.stdout.write(self.style.ERROR('변환되지 않은 인접지역 데이터: %s' % val))

        return connected_count
