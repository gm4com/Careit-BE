import json
import logging

from openpyxl import Workbook

from django.shortcuts import render, HttpResponseRedirect
from django.conf import settings
from django.views.generic import ListView
from django.utils import timezone
from django.http import HttpResponse
from django.utils.http import urlquote
from django.contrib import admin
from django.contrib.admin.views.main import ChangeList

from drf_yasg.utils import swagger_auto_schema
from drf_yasg.openapi import Schema
from rest_framework import permissions, response, viewsets, mixins, serializers

from common.utils import ServerScript, SlackWebhook


logger = logging.getLogger('django')


"""
DRF API Permission
"""


class IsSuperUser(permissions.BasePermission):
    """
    슈퍼유져만 허용하는 권한
    """

    def has_permission(self, request, view):
        return request.user.is_superuser


class IsLocalhost(permissions.BasePermission):
    """
    로컬에서만 허용하는 권한
    """

    def has_permission(self, request, view):
        return request.META['REMOTE_ADDR'] in settings.INTERNAL_IPS


"""
Errors
"""


def permission_denied(request, exception):
    return render(request, '403.html', status=403)


def page_not_found(request, exception):
    return render(request, '404.html', status=404)


def server_error(request):
    return render(request, '500.html', status=500)


"""
Decorators
"""


def swagger_auto_boolean_schema(**kwargs):
    boolean_schema = {
        200: Schema(
            type='object',
            properties={'result': {'type': 'boolean', 'title': '처리 결과'}},
            title='BooleanResult'
        ),
    }
    if 'responses' in kwargs:
        kwargs['responses'].update(boolean_schema)
    else:
        kwargs.update({
            'responses': boolean_schema
        })
    return swagger_auto_schema(**kwargs)


"""
Webhook Callback Views
"""


class WebhookCallbackSerializer(serializers.Serializer):
    pass


class WebhookCallbackView(mixins.CreateModelMixin, viewsets.GenericViewSet):
    """
    웹훅 뷰
    """
    serializer_class = WebhookCallbackSerializer
    permission_classes = (permissions.AllowAny,)
    refs = {
        'refs/heads/staging': 'Staging',
        'refs/heads/master': 'Production',
    }

    def create(self, request, *args, **kwargs):
        server_name = settings.SERVER_DEPLOY_NAME.capitalize()
        payload = ''
        try:
            payload = json.loads(request.data['payload'])
            target = self.refs[payload['ref']]
            name = payload['pusher']['name']
        except:
            logger.error(payload)
        else:
            msg = '%s 님이 %s 서버를 업데이트 했습니다.' % (name, target)
            logger.warning(msg)
            if target == server_name:
                slack = SlackWebhook().channel('anyman__80dev')
                slack.script_msg(msg, result=ServerScript().deploy())
            else:
                logger.warning('채널이 지정되지 않아서 알림을 보낼 수 없습니다.')
        return response.Response({})


"""
Excel Export Views
"""


class ModelExportBaseView(ListView):
    """
    엑셀 내보내기 기본 뷰
    """
    file_type = 'xlsx'
    file_content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    data_handler = 'single_sheet'
    columns = ()
    header_row = 1
    header_col = 1
    data_row = 2
    data_col = 1
    workbook = None
    worksheet = None

    def dispatch(self, request, *args, **kwargs):
        if 'queryset' in kwargs:
            self.queryset = kwargs['queryset']
        return super(ModelExportBaseView, self).dispatch(request, *args, **kwargs)

    def get_queryset(self):
        if self.queryset:
            return self.queryset
        return super(ModelExportBaseView, self).get_queryset()

    def get(self, request, *args, **kwargs):
        response = HttpResponse(
            content_type=self.file_content_type,
        )
        response['Content-Disposition'] = 'attachment; filename=%s' % urlquote(self.get_filename())
        response_function = getattr(self, 'get_%s_response' % self.data_handler)
        return response_function(response)

    def get_filename(self):
        return '%s_%s.%s' % (self.get_model_name(), timezone.now().strftime('%Y-%m-%d'), self.file_type)

    def get_model_name(self):
        return self.model._meta.verbose_name

    def initialize(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = str(self.get_model_name())

    def get_single_sheet_response(self, response):
        # initialize
        self.initialize()

        # header
        self.handle_headers()

        # data
        self.handle_queryset(self.get_queryset())

        self.workbook.save(response)
        return response

    def handle_headers(self):
        for col_num, column in enumerate(self.columns, self.header_col):
            cell = self.worksheet.cell(row=self.header_row, column=col_num)
            cell.value = column[1]

    def handle_queryset(self, queryset, start_row=None):
        row = start_row or self.data_row
        for obj in queryset:
            row = self.handle_query(obj, row)
        return row

    def handle_query(self, obj, row):
        col = self.data_col
        for field_name, _ in self.columns:
            cell = self.worksheet.cell(row=row, column=col)
            if '__' in field_name:
                objects = [obj]
                for field in field_name.split('__'):
                    value = getattr(objects[-1], field, None)
                    objects.append(value)
            else:
                value = getattr(obj, field_name, None)
            if value is None:
                handler = getattr(self, 'get_field_' + field_name, '')
                if callable(handler):
                    value = handler(obj)
            elif callable(value):
                value = value()
            cell.value = str(value) if value is not None else ''
            col += 1
        return row + 1

    def get_field_empty(self, obj):
        return ''


class FilteredQuerySetChangeList(ChangeList):
    def get_queryset(self, request):
        qs = super(FilteredQuerySetChangeList, self).get_queryset(request)
        self.model_admin.filtered_queryset = qs
        return qs


class ExcelFilter(admin.SimpleListFilter):
    """
    엑셀 다운로드 필터
    """
    title = '엑셀 다운로드'
    parameter_name = 'download'

    def lookups(self, request, model_admin):
        return ()

    def queryset(self, request, queryset):
        return queryset


class FilteredExcelDownloadMixin:
    excel_download_view = None

    def get_list_filter(self, request):
        filters = super(FilteredExcelDownloadMixin, self).get_list_filter(request)
        return list(filters) + [ExcelFilter]

    def get_changelist(self, request, **kwargs):
        return FilteredQuerySetChangeList

    def changelist_view(self, request, extra_context=None):
        view = super(FilteredExcelDownloadMixin, self).changelist_view(request, extra_context=extra_context)
        if request.GET.get('download'):
            return self.excel_download_view.as_view()(request, queryset=self.filtered_queryset)
        return view
